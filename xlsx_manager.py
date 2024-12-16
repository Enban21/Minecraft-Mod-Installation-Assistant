# xlsx_manager.py
import openpyxl

def load_xlsx(excel_path, app):
    try:
        workbook = openpyxl.load_workbook(excel_path)
        mod_categories = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            mods = []
            
            # A1セルをフォルダ階層とする
            folder = sheet['A1'].value or ''  # 無記載なら空文字
            
            # A2以降の行からMOD情報を取得
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # MODファイル名が空でない場合
                    mod = {
                        "folder": folder,
                        "file": row[0],
                        "link": row[1] if len(row) > 1 else '',
                        "description": row[2] if len(row) > 2 else '',
                    }
                    mods.append(mod)
            
            mod_categories[sheet_name] = mods
        
        app.mod_categories = mod_categories
        app.update_list_widget()

    except Exception as e:
        raise Exception(f'Excelファイルの読み込みに失敗しました: {e}')
